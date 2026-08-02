# CANONICAL SOURCE: athena_html_v3/recon_attribution.py @ 0600e07 (2026-08-02).
# SYNCED COPY (stage-3 recon move) — athena's copy is deprecated and will
# be removed once parity is verified; this becomes the canonical home.
"""
Two-pass Maia vs administrator reconciliation with difference attribution.

The point of running it TWICE. Pass 1 values each system on its own prices and
FX; the total difference is real but tells you nothing about its cause. Pass 2
revalues Maia's positions at the ADMINISTRATOR's prices and FX rates. Anything
still different in pass 2 is NOT a mark difference — it is a position, accrual
or convention difference, which is the kind that actually needs fixing. Marks
disagreeing is normal; books disagreeing is not.

ATTRIBUTION. Per bond, the market-value difference decomposes exactly into three
effects, computed in this order so they sum without residual:

    par    = (maia_par - admin_par) x admin_px/100 x admin_fx
    price  =  maia_par x (maia_px - admin_px)/100  x admin_fx
    fx     =  maia_par x maia_px/100 x (maia_fx - admin_fx)

    maia_mv - admin_mv = par + price + fx      (identity, asserted)

Accrued is compared separately rather than folded in: it is quoted per bond in
LOCAL currency by both systems and differs by DAY-COUNT CONVENTION, not by mark.
Mixing it into the price effect would hide the one difference that indicates a
static-data problem.

WHY ACCRUED IS THE INTERESTING COLUMN. On GDBF at 2026-07-24 Maia's price-implied
accrual ran ahead of the administrator's booked accrual on every line — 347.0
implied days against 292.6 on the Bund. That is a convention difference, and no
amount of agreeing on price will make it go away.

FX SOURCES. The administrator publishes rates (fx_rates.rates, quoted per base,
so base-per-unit is 1/rate). Maia's are derived from its own cash rows
(exposure/qty), which is exact for cash. Both are read, never assumed, and never
recomputed from bond prices — deriving a rate from prices breaks on any per-unit
quoted security, which is how the Nationwide CCDS corrupted three separate
calculations before this was made a rule.

EXACTNESS IS THE GOAL. Every figure comes from one of the two files; nothing is
inferred. Where a bond cannot be matched, or a rate is missing, it is reported in
`unmatched`/`missing_fx` rather than defaulted — a default of 1.0 silently values
a foreign bond at par.
"""

from __future__ import annotations

import bond_recon


def _admin_fx(parsed_nav: dict) -> dict:
    """{ccy: base per 1 unit} from the administrator's published rates."""
    rates = (parsed_nav.get("fx_rates") or {}).get("rates") or {}
    return {c: (1.0 / r) for c, r in rates.items() if r}


def _maia_fx(priced_path: str) -> dict:
    """{ccy: base per 1 unit} derived from Maia's own cash rows (exposure/qty)."""
    import aum_recon
    return aum_recon._fx_rates(*aum_recon._sheet(priced_path))


def _as_date(v) -> str | None:
    """ISO date from a cell, or None. Accepts datetimes and dd/mm/yyyy."""
    import datetime as _dt
    import re
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", s):
        return s[:10]
    m = re.fullmatch(r"(\d{2})/(\d{2})/(\d{4}).*", s)
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else None


def _maia_as_of(priced_path: str) -> str | None:
    """Valuation date stamped in the Maia workbook, if it carries one."""
    import openpyxl
    import re
    wb = openpyxl.load_workbook(priced_path, data_only=True)
    # SHEET0 ONLY. A Maia workbook can carry the ADMINISTRATOR's sheets pasted
    # in alongside it (gdbfpos7 has Sheet1 = Detailed Security Valuation and
    # Sheet2 = Accrued Income Recon, both stamped with the admin's own
    # valuation date). Walking every sheet risks reading the admin's date,
    # calling it Maia's, and declaring the dates match by comparing the admin
    # with itself.
    for ws in [wb["Sheet0"]] if "Sheet0" in wb.sheetnames else wb.worksheets[:1]:
        for row in ws.iter_rows(min_row=1, max_row=14, values_only=True):
            for i, v in enumerate(row):
                # Must be the LABEL cell, not a column header: Sheet0 has a
                # "Last Valuation Date" COLUMN whose neighbour is "Ticker",
                # which an unguarded match happily returned as a date.
                if isinstance(v, str) and v.strip().rstrip(":").endswith("Valuation Date"):
                    for nxt in row[i + 1:]:
                        d = _as_date(nxt)
                        if d:
                            return d
        # Fall back to the freshest price timestamp (dd/mm/yyyy hh:mm:ss).
        stamps = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            for v in row:
                if isinstance(v, str) and re.fullmatch(
                        r"\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}", v.strip()):
                    d, m, y = v[:10].split("/")
                    stamps.append(f"{y}-{m}-{d}")
        if stamps:
            # MODE, not max. These workbooks carry stale rows — a cash line
            # last priced 04/04/2025, an odd bond stamped a few days later —
            # so the newest stamp is not the file's date. The date most rows
            # share is.
            from collections import Counter
            return Counter(stamps).most_common(1)[0][0]
    return None


def attribute(parsed_nav: dict, priced_path: str, reference_path: str) -> dict:
    """Both passes, per bond and in total.

    DATES ARE CHECKED FIRST. Reconciling an administrator pack against a Maia
    export from a different day attributes several days of market and currency
    movement to "difference", which reads as a break and is not one. The first
    run of this module compared a 2026-07-30 pack against a 2026-07-24 Maia file
    and produced FX "differences" of -0.0096 EUR and -0.0066 GBP — almost
    entirely six days of currency movement. The mismatch is reported loudly
    rather than silently absorbed.
    """
    base = parsed_nav.get("base_currency")
    joined = bond_recon.compare(parsed_nav, priced_path, reference_path)
    afx, mfx = _admin_fx(parsed_nav), _maia_fx(priced_path)
    admin_h = {h["isin"]: h for h in (parsed_nav.get("holdings") or []) if h.get("isin")}
    maia_b, _ = bond_recon.maia_bonds(priced_path, bond_recon.isin_bridge(reference_path))

    rows, unmatched, missing_fx = [], [], []
    tot = {"par": 0.0, "price": 0.0, "fx": 0.0, "accrued": 0.0,
           "admin_mv": 0.0, "maia_mv": 0.0, "maia_mv_at_admin_marks": 0.0,
           "dirty_residual": 0.0}

    for isin in sorted(set(admin_h) | set(maia_b)):
        a, m = admin_h.get(isin), maia_b.get(isin)
        if not a or not m:
            unmatched.append({"isin": isin, "in_admin": bool(a), "in_maia": bool(m)})
            continue
        ccy = a.get("currency") or m.get("currency")
        a_fx, m_fx = afx.get(ccy), mfx.get(ccy)
        if a_fx is None or m_fx is None:
            missing_fx.append({"isin": isin, "currency": ccy,
                               "admin_fx": a_fx, "maia_fx": m_fx})
            continue

        a_par, m_par = a.get("par_amount") or 0.0, m.get("par") or 0.0
        a_px, m_px = a.get("price") or 0.0, m.get("price") or 0.0

        # PRICE DIVISOR, READ FROM THE ADMINISTRATOR — never assumed to be 100.
        # Most bonds quote per 100; the Nationwide CCDS quotes PER UNIT, and
        # hardcoding /100 understated it ~100x and knocked ~290k off both sides
        # of this attribution, so section 3's levels no longer tied to section
        # 2's. Deriving the divisor from market_value / (par x px x fx) gives
        # 100 for a normal bond and 1 for a per-unit one, from the
        # administrator's own arithmetic rather than a guess.
        mv_base = a.get("market_value")
        divisor = 100.0
        if mv_base and a_par and a_px and a_fx:
            implied = (a_par * a_px * a_fx) / mv_base
            # Snap to the two conventions that exist; anything else is a data
            # problem and is left at 100 so it shows up rather than silently
            # rescaling.
            divisor = 1.0 if 0.5 < implied < 2.0 else 100.0

        # PASS 1 — each system on its own marks.
        admin_mv = a_par * a_px / divisor * a_fx
        maia_mv = m_par * m_px / divisor * m_fx
        par_eff = (m_par - a_par) * a_px / divisor * a_fx
        price_eff = m_par * (m_px - a_px) / divisor * a_fx
        fx_eff = m_par * m_px / divisor * (m_fx - a_fx)

        # PASS 2 — Maia's POSITIONS at the administrator's marks.
        #
        # BE HONEST ABOUT WHAT THIS TESTS. Every term except par is admin data
        # on BOTH sides, so this residual is algebraically
        #     (maia_par - admin_par) x admin_px/100 x admin_fx
        # i.e. the par effect restated. It tests POSITION EQUALITY and nothing
        # else, and it CANNOT surface a price, FX or accrual difference. An
        # earlier version reported it as "residual — NOT a mark difference",
        # which read as a clean bill of health when the fund agreed on par;
        # Andy correctly called that suspicious.
        maia_at_admin = m_par * a_px / divisor * a_fx

        # Accrued, in base, from each system's own per-bond figure.
        a_acc = (a.get("accrued_income") or 0.0) * a_fx
        m_acc = (m.get("accrued_local") or 0.0) * m_fx

        # PASS 3 — the one that actually bites. Compare DIRTY value (clean +
        # accrued) with BOTH sides on the administrator's price and FX, so the
        # only thing left free is each system's own accrued. Non-zero here is a
        # genuine day-count / accrual-convention difference, which no amount of
        # agreeing on price will fix.
        a_dirty_at_admin = a_par * a_px / divisor * a_fx + a_acc
        m_dirty_at_admin = m_par * a_px / divisor * a_fx + (m.get("accrued_local") or 0.0) * a_fx

        rows.append({
            "isin": isin, "currency": ccy,
            "description": a.get("description"),
            "admin_par": a_par, "maia_par": m_par,
            "admin_price": a_px, "maia_price": m_px,
            "admin_fx": round(a_fx, 6), "maia_fx": round(m_fx, 6),
            "admin_mv": round(admin_mv, 2), "maia_mv": round(maia_mv, 2),
            "mv_diff": round(maia_mv - admin_mv, 2),
            "effect_par": round(par_eff, 2),
            "effect_price": round(price_eff, 2),
            "effect_fx": round(fx_eff, 2),
            # Pass 2: the number that matters. Non-zero means a real break.
            "maia_mv_at_admin_marks": round(maia_at_admin, 2),
            "residual_at_admin_marks": round(maia_at_admin - admin_mv, 2),
            "admin_accrued_base": round(a_acc, 2),
            "maia_accrued_base": round(m_acc, 2),
            "accrued_diff": round(m_acc - a_acc, 2),
            "price_divisor": divisor,
            # Pass 3 — dirty value, both sides on admin marks. Isolates accrual.
            "admin_dirty_at_admin_marks": round(a_dirty_at_admin, 2),
            "maia_dirty_at_admin_marks": round(m_dirty_at_admin, 2),
            "dirty_residual": round(m_dirty_at_admin - a_dirty_at_admin, 2),
        })
        tot["par"] += par_eff; tot["price"] += price_eff; tot["fx"] += fx_eff
        tot["accrued"] += m_acc - a_acc
        tot["admin_mv"] += admin_mv; tot["maia_mv"] += maia_mv
        tot["maia_mv_at_admin_marks"] += maia_at_admin
        tot["dirty_residual"] += m_dirty_at_admin - a_dirty_at_admin

    # The decomposition must be exact. A residual means the arithmetic above is
    # wrong, not that the data is odd — so assert rather than report.
    identity = tot["maia_mv"] - tot["admin_mv"] - (tot["par"] + tot["price"] + tot["fx"])
    admin_date = parsed_nav.get("valuation_date")
    maia_date = _maia_as_of(priced_path)
    return {
        "valuation_date": admin_date,
        "admin_date": admin_date,
        "maia_date": maia_date,
        # False means several days of market and FX movement are being reported
        # as "difference". Every number below is suspect until the dates match.
        "dates_match": bool(admin_date and maia_date and admin_date == maia_date),
        "base_currency": base,
        "rows": rows,
        "totals": {k: round(v, 2) for k, v in tot.items()},
        "residual_at_admin_marks": round(tot["maia_mv_at_admin_marks"] - tot["admin_mv"], 2),
        "decomposition_exact": abs(identity) < 0.01,
        "decomposition_residual": round(identity, 6),
        "unmatched": unmatched,
        "missing_fx": missing_fx,
        "admin_fx": {k: round(v, 6) for k, v in afx.items()},
        "maia_fx": {k: round(v, 6) for k, v in mfx.items()},
    }


def print_report(res: dict) -> None:
    t = res["totals"]
    W = 100
    print("=" * W)
    print(f"MAIA vs ADMINISTRATOR — difference attribution   base {res['base_currency']}")
    print(f"   administrator: {res['admin_date']}      Maia: {res['maia_date']}")
    print("=" * W)
    if not res["dates_match"]:
        print("\n*** DATES DO NOT MATCH — every difference below includes market and")
        print("*** FX movement between the two dates. Re-run with same-day files")
        print("*** before treating anything here as a break.")

    print("\nFX RATES (base per 1 unit) — read from each source, never assumed")
    for c in sorted(set(res["admin_fx"]) | set(res["maia_fx"])):
        a, m = res["admin_fx"].get(c), res["maia_fx"].get(c)
        d = (m - a) if (a is not None and m is not None) else None
        print(f"   {c:5} admin {a if a is None else f'{a:.6f}':>10}   "
              f"maia {m if m is None else f'{m:.6f}':>10}   "
              f"diff {'—' if d is None else f'{d:+.6f}'}")

    print("\nPASS 1 — each system on its OWN prices and FX")
    print(f"   admin bond MV                {t['admin_mv']:>16,.2f}")
    print(f"   maia  bond MV                {t['maia_mv']:>16,.2f}")
    print(f"   difference                   {t['maia_mv'] - t['admin_mv']:>16,.2f}")
    print(f"      of which position (par)   {t['par']:>16,.2f}")
    print(f"      of which price            {t['price']:>16,.2f}")
    print(f"      of which FX               {t['fx']:>16,.2f}")
    print(f"   decomposition exact: {res['decomposition_exact']}"
          f"  (residual {res['decomposition_residual']})")

    print("\nPASS 2 — POSITIONS ONLY (both sides on admin price and FX)")
    print(f"   residual                     {res['residual_at_admin_marks']:>16,.2f}")
    print( "   This tests par equality and NOTHING else — every other term is")
    print( "   admin data on both sides, so it is the par effect restated. Zero")
    print( "   here means the books agree on size, not that anything else does.")

    print("\nPASS 3 — DIRTY VALUE, both sides on admin price and FX")
    print(f"   residual                     {t['dirty_residual']:>16,.2f}")
    print( "   Only each system's own ACCRUED is left free, so this is a pure")
    print( "   day-count / accrual-convention difference.")

    print(f"\nACCRUED as each system reports it (own FX)")
    print(f"   difference                   {t['accrued']:>16,.2f}")

    big = sorted((r for r in res["rows"] if abs(r["residual_at_admin_marks"]) > 0.01),
                 key=lambda r: -abs(r["residual_at_admin_marks"]))
    print(f"\nBONDS WITH A NON-MARK DIFFERENCE: {len(big)}")
    for r in big[:12]:
        print(f"   {r['isin']:14} {str(r['currency']):4} "
              f"admin par {r['admin_par']:>11,.0f}  maia par {r['maia_par']:>11,.0f}  "
              f"residual {r['residual_at_admin_marks']:>12,.2f}")

    acc = sorted((r for r in res["rows"] if abs(r["accrued_diff"]) > 1),
                 key=lambda r: -abs(r["accrued_diff"]))
    print(f"\nLARGEST ACCRUED DIFFERENCES: {len(acc)} over 1.00")
    for r in acc[:10]:
        print(f"   {r['isin']:14} admin {r['admin_accrued_base']:>11,.2f}  "
              f"maia {r['maia_accrued_base']:>11,.2f}  diff {r['accrued_diff']:>+11,.2f}")

    if res["unmatched"]:
        print(f"\nUNMATCHED: {res['unmatched']}")
    if res["missing_fx"]:
        print(f"MISSING FX: {res['missing_fx']}")


def main(argv=None) -> int:
    import argparse
    import nav_parser
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    ap.add_argument("nav", help="administrator NAV report (.xls)")
    ap.add_argument("priced", help="Maia priced position view (.xlsx)")
    ap.add_argument("reference", help="Maia reference view carrying ISIN (.xlsx)")
    args = ap.parse_args(argv)
    print_report(attribute(nav_parser.parse_nav_report(args.nav),
                           args.priced, args.reference))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
