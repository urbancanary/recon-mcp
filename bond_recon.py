# CANONICAL SOURCE: athena_html_v3/bond_recon.py @ 0600e07 (2026-08-02).
# SYNCED COPY (stage-3 recon move) — athena's copy is deprecated and will
# be removed once parity is verified; this becomes the canonical home.
"""
Per-bond reconciliation: Maia vs the administrator, from FILES.

No copy-paste. The older four-source recon (widgets/nav-recon-widget.js,
_renderGcrifFromDataset) gets its Maia side by pasting TSV into the page with a
hardcoded CNH rate — a GCRIF-era path. This reads both sides from the files we
already ingest: the Waystone .xls via nav_parser, and Maia's position-view .xlsx
directly.

THE ISIN PROBLEM, AND THE BRIDGE THAT SOLVES IT. Maia's PRICED position view
(gdbfpos7-style, 34 columns: Last Px, Dirty price, Accrued Total, Qty, Exposure)
carries NO ISIN — only Ticker and Description. Joining on anything else is
unsafe: an attempt to match on (quantity, currency) silently collapsed six GDBF
bonds that all hold 220,000 GBP onto one row and produced differences up to
722bp that were pure join error.

But Maia CAN export ISIN — a second view (gdbfpos5-style, 42 columns: ISIN,
Ticker, Issuer, Currency, Asset Class, Country of Risk, Rho, Duration) carries
it for all 28 bonds. So the join goes:

    priced view --Ticker--> reference view --ISIN--> administrator holdings

Both Maia views must be from the same snapshot or the bridge maps stale tickers.
`unresolved` names any ticker the bridge could not carry, and those bonds are
excluded rather than guessed at.

A STALE BRIDGE DOES NOT FAIL LOUDLY — IT REPORTS ABSENCE. This bit us on
2026-07-31. The bridge was built from a 24 July reference view, so four bonds
bought since had no entry, and the report called them "missing from Maia" and
produced a sum-of-differences of -2,881,558.00. They were not missing; the
bridge simply predated them. Three of them turned out to be live duration
switches (the fund sold the 2046/2053s and bought the 2033/34/36s between the
administrator's 12:00 strike and Maia's 14:51 snapshot) — a real and interesting
finding that the stale bridge had disguised as missing data.

So the dates are now checked and carried in the result: `bridge_stale` is True
whenever the reference view's date differs from the priced view's, and when it
is True the position-break and coverage sections are marked unreliable rather
than being presented as findings. Absence of evidence is not evidence of a
break.

ASK MAIA FOR ONE EXPORT CARRYING BOTH. The bridge works but is fragile — it
needs two files kept in step. The reference view proves the field exists, so
adding ISIN to the priced view is a request, not a feature request.

WHAT IS COMPARED. Price, par and market value per bond, plus the price
difference in basis points, which is the number that tells you whether two
systems actually disagree about a security. Measured on GDBF 2026-07-24:
25 of 28 bonds differ by more than 10bp, the largest being Intel
(US458140CK47) at +294bp and SW Finance (XS2731297235) at -168bp.
"""

from __future__ import annotations

import openpyxl

# Columns are resolved BY NAME via aum_recon's shared resolver, so every Maia
# export shape works and a layout change cannot silently shift a field. This
# module previously carried its own fixed indices for the 34-column view, which
# meant the 13-column allocation view and the 629-column full export both read
# as "no bonds".
from aum_recon import _Cols, _classify, _sheet          # noqa: E402


def _num(v):
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return None


def isin_bridge(reference_path: str) -> dict:
    """{Ticker: ISIN} from a Maia reference view that carries ISIN.

    Use `isin_bridge_meta` when the caller needs the snapshot date too.
    """
    return isin_bridge_meta(reference_path)["map"]


def isin_bridge_meta(reference_path: str) -> dict:
    """{"map": {Ticker: ISIN}, "as_of": date|None, "path": str, "count": int}."""
    ws = openpyxl.load_workbook(reference_path, data_only=True)["Sheet0"]
    hdr = [str(c).strip() if c is not None else ""
           for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    if "ISIN" not in hdr or "Ticker" not in hdr:
        raise ValueError(f"{reference_path} has no ISIN/Ticker columns — "
                         "this is not a Maia reference view")
    i_isin, i_tkr = hdr.index("ISIN"), hdr.index("Ticker")
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        isin, tkr = row[i_isin], row[i_tkr]
        if isin and tkr:
            out[str(tkr).strip()] = str(isin).strip()

    from recon_attribution import _maia_as_of
    try:
        as_of = _maia_as_of(reference_path)
    except Exception:
        as_of = None
    return {"map": out, "as_of": as_of, "path": reference_path, "count": len(out)}


def maia_bonds(priced_path: str, bridge: dict) -> tuple[dict, list]:
    """({isin: bond}, unresolved_tickers) from any Maia position view.

    Where the export carries ISIN itself (the 629-column full export does), that
    ISIN is used directly and the bridge is bypassed — no bridge, no staleness.
    """
    cols, rows = _sheet(priced_path)
    bonds, unresolved = {}, []
    for row in rows:
        if _classify(cols, row) != "bond":
            continue
        qty = _num(cols.get(row, "qty"))
        if qty is None:          # subtotal rows carry no quantity
            continue
        ticker = str(cols.get(row, "ticker") or "").strip()
        isin = cols.get(row, "isin") or bridge.get(ticker)
        if not isin:
            # Keep the VALUE, not just the name. An unidentified position is
            # still a position: its value can be compared in aggregate against
            # the administrator's unmatched value without knowing which bond is
            # which. Dropping it entirely threw away the strongest evidence
            # available whenever the bridge was stale.
            unresolved.append({
                "ticker": ticker,
                "description": cols.get(row, "description"),
                "qty": qty,
                "exposure_base": _num(cols.get(row, "exposure")),
                "currency": cols.get(row, "currency"),
            })
            continue
        bonds[str(isin).strip()] = {
            "ticker": ticker,
            "description": cols.get(row, "description"),
            "currency": cols.get(row, "currency"),
            "par": qty,
            "price": _num(cols.get(row, "last_px")),
            "dirty_price": _num(cols.get(row, "dirty_px")),
            "accrued_local": _num(cols.get(row, "accrued_local")),
            "exposure_base": _num(cols.get(row, "exposure")),
        }
    return bonds, unresolved


def compare(parsed_nav: dict, priced_path: str, reference_path: str) -> dict:
    """Join Maia's priced view to the administrator's holdings on ISIN."""
    from recon_attribution import _maia_as_of
    meta = isin_bridge_meta(reference_path)
    bridge = meta["map"]
    maia, unresolved = maia_bonds(priced_path, bridge)

    try:
        priced_as_of = _maia_as_of(priced_path)
    except Exception:
        priced_as_of = None
    bridge_stale = bool(meta["as_of"] and priced_as_of
                        and meta["as_of"] != priced_as_of)
    admin = {h["isin"]: h for h in (parsed_nav.get("holdings") or []) if h.get("isin")}

    rows = []
    for isin in sorted(set(maia) | set(admin)):
        a, m = admin.get(isin), maia.get(isin)
        a_px = (a or {}).get("price")
        m_px = (m or {}).get("price")
        px_diff = None if (a_px is None or m_px is None) else round(m_px - a_px, 6)
        bp = (None if (px_diff is None or not a_px)
              else round(px_diff / a_px * 10000, 1))
        a_par, m_par = (a or {}).get("par_amount"), (m or {}).get("par")
        rows.append({
            "isin": isin,
            "description": (a or m or {}).get("description"),
            "currency": (a or {}).get("currency") or (m or {}).get("currency"),
            "in_admin": a is not None,
            "in_maia": m is not None,
            "admin_par": a_par,
            "maia_par": m_par,
            # A par break is a POSITION break — a different and more serious
            # thing than a price difference, so it is flagged separately.
            "par_diff": (None if (a_par is None or m_par is None)
                         else round(m_par - a_par, 2)),
            "admin_price": a_px,
            "maia_price": m_px,
            "price_diff": px_diff,
            "price_diff_bp": bp,
            "admin_mv_base": (a or {}).get("market_value"),
            "admin_accrued_base": (a or {}).get("accrued_income"),
            # DIRTY, to match Maia. Maia's Exposure is dirty (clean + accrued);
            # the administrator states clean market value and accrued
            # separately. Comparing admin CLEAN against Maia DIRTY understates
            # the administrator by the whole accrued balance — on the GDBF
            # 2026-07-31 unmatched positions that was 37,650 on 2.85m.
            "admin_dirty_base": (None if (a or {}).get("market_value") is None
                                 else round(((a or {}).get("market_value") or 0.0)
                                            + ((a or {}).get("accrued_income") or 0.0), 2)),
            "maia_exposure_base": (m or {}).get("exposure_base"),
        })

    priced = [r for r in rows if r["price_diff_bp"] is not None]

    # With a stale bridge, an unresolved ticker means "the bridge predates this
    # bond", not "Maia does not hold it" — so coverage and position breaks
    # cannot be reported as findings. The price comparison on the bonds that DID
    # resolve is still valid: those joined on a real, matching ISIN.
    warnings = []
    if bridge_stale:
        warnings.append(
            f"ISIN bridge is from {meta['as_of']} but the priced view is "
            f"{priced_as_of}. Bonds traded between those dates have no bridge "
            f"entry and will look missing when they are not. Position breaks "
            f"and coverage below are NOT reliable; supply a reference view "
            f"dated {priced_as_of}.")
    elif unresolved and not meta["as_of"]:
        warnings.append(
            f"{len(unresolved)} ticker(s) unresolved and the bridge carries no "
            f"date, so staleness cannot be ruled out as the cause.")

    return {
        "valuation_date": parsed_nav.get("valuation_date"),
        "bridge": {"path": meta["path"], "as_of": meta["as_of"],
                   "entries": meta["count"]},
        "priced_as_of": priced_as_of,
        "bridge_stale": bridge_stale,
        # Consumers must gate the position-break and coverage sections on this.
        "coverage_reliable": not bridge_stale,
        "warnings": warnings,
        "rows": rows,
        "matched": len(priced),
        "admin_only": [r["isin"] for r in rows if not r["in_maia"]],
        "maia_only": [r["isin"] for r in rows if not r["in_admin"]],
        "unresolved_tickers": [u["ticker"] for u in unresolved],
        "unresolved_positions": unresolved,
        "unresolved_value": round(
            sum(u.get("exposure_base") or 0.0 for u in unresolved), 2),
        "par_breaks": [r for r in rows if r["par_diff"] not in (None, 0)],
        "over_10bp": sorted((r for r in priced if abs(r["price_diff_bp"]) > 10),
                            key=lambda r: -abs(r["price_diff_bp"])),
    }


def print_report(res: dict) -> None:
    print(f"PER-BOND RECON — Maia vs administrator   {res['valuation_date']}\n")
    for w in res.get("warnings") or []:
        print(f"  !! {w}\n")
    print(f"  {'ISIN':14} {'ccy':4} {'par':>11} {'admin px':>10} "
          f"{'maia px':>10} {'diff':>9} {'bp':>8}")
    for r in res["rows"]:
        if r["price_diff_bp"] is None:
            flag = "admin only" if not r["in_maia"] else "maia only"
            print(f"  {r['isin']:14} {str(r['currency'] or ''):4} "
                  f"{'':>11} {'':>10} {'':>10} {'':>9} {flag:>8}")
            continue
        mark = "  <<<" if abs(r["price_diff_bp"]) > 100 else ""
        print(f"  {r['isin']:14} {str(r['currency'] or ''):4} {r['admin_par']:>11,.0f} "
              f"{r['admin_price']:>10.4f} {r['maia_price']:>10.4f} "
              f"{r['price_diff']:>9.4f} {r['price_diff_bp']:>8.1f}{mark}")
    print(f"\n  matched {res['matched']}   >10bp {len(res['over_10bp'])}   "
          f"par breaks {len(res['par_breaks'])}")
    if not res.get("coverage_reliable", True):
        print("  coverage (admin-only / Maia-only / position breaks) SUPPRESSED "
              "— stale ISIN bridge, see warning above")
    else:
        if res["admin_only"]:
            print(f"  in admin only: {res['admin_only']}")
        if res["maia_only"]:
            print(f"  in Maia only:  {res['maia_only']}")
    if res["unresolved_tickers"]:
        print(f"  tickers the ISIN bridge could not resolve: {res['unresolved_tickers']}")
    if res["par_breaks"] and res.get("coverage_reliable", True):
        print("\n  POSITION BREAKS (par differs — more serious than a price gap):")
        for r in res["par_breaks"]:
            print(f"    {r['isin']} admin {r['admin_par']:,.0f} vs "
                  f"maia {r['maia_par']:,.0f}  ({r['par_diff']:+,.0f})")


def main(argv=None) -> int:
    import argparse
    import nav_parser
    ap = argparse.ArgumentParser(description="Per-bond Maia vs administrator recon")
    ap.add_argument("nav", help="administrator NAV report (.xls)")
    ap.add_argument("priced", help="Maia priced position view (.xlsx)")
    ap.add_argument("reference", help="Maia reference view carrying ISIN (.xlsx)")
    args = ap.parse_args(argv)
    print_report(compare(nav_parser.parse_nav_report(args.nav),
                         args.priced, args.reference))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
