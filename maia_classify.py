"""Content-based classification of files landing from the Guinness OneDrive
MAIA folder (or any upload). Filenames there carry NO reliable signal —
`tempMaia45782….xlsx`, `maia postions gdbf.xlsx`, `gdbfptc…` — so every file
is identified by READING it: sheet names first, then Sheet0's header row.

Types (survey of the real folder, 2026-08-03/04):

    admin_pack        Waystone NAV report (sheets: Balance_Sheet,
                      Share_Class_Price_Report, OpenCurrency…) — sometimes
                      saved as .xlsx, not just the mailed .xls
    maia_priced       34-col priced view: Qty + Exposure + Accrued Total —
                      the AUM recon's preferred food
    maia_allocation   13-col allocation view: Grouping + Holding + Exp (Fund CCY)
    maia_full         629-col full export: carries ISIN (bridge view)
    maia_positions    NEW positions/exposure shape: Ticker/Qty/Strategy/
                      Instrument ID/Exp Fund CCY — no Grouping, no accrued
    maia_aum          fund-level AUM summary: AUM ($) / Calculated NAV per date
    maia_compliance   pre-trade compliance rules: Rule ID / Fail Msg / Rule Type
    unknown           none of the above (decks, docs, shortcuts…)

Detection is deliberately ORDERED most-specific-first, and every rule keys on
columns that define what the file can ANSWER, mirroring aum_recon's
capability philosophy.
"""

from __future__ import annotations

import io

import openpyxl

# Sheet names that identify a Waystone/InvestOne administrator pack.
_ADMIN_SHEETS = {"balance_sheet", "share_class_price_report", "opencurrency",
                 "detailed_security_valuation"}


def _norm(h) -> str:
    return " ".join(str(h or "").split()).strip().lower()


def _sheet0_headers(wb) -> set[str]:
    name = next((s for s in wb.sheetnames if s.lower().startswith("sheet0")), None)
    if not name:
        name = wb.sheetnames[0]
    ws = wb[name]
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return {_norm(c) for c in row if c is not None}


def classify_xlsx(file_bytes: bytes) -> dict:
    """{"type": <one of the types above>, "detail": str} — never raises for
    a readable workbook; unreadable input classifies as unknown."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True,
                                    data_only=True)
    except Exception as e:
        return {"type": "unknown", "detail": f"not a readable workbook: {e}"}

    sheets = {s.lower() for s in wb.sheetnames}
    if sheets & _ADMIN_SHEETS:
        return {"type": "admin_pack",
                "detail": f"administrator sheets present: {sorted(sheets & _ADMIN_SHEETS)}"}

    try:
        hdr = _sheet0_headers(wb)
    except Exception as e:
        return {"type": "unknown", "detail": f"no readable header row: {e}"}

    def has(*names):
        return all(_norm(n) in hdr for n in names)

    if has("rule id", "fail msg"):
        return {"type": "maia_compliance", "detail": "pre-trade compliance rules"}
    if has("aum", "calculated nav") or has("aum ($)", "calculated nav"):
        return {"type": "maia_aum", "detail": "fund-level AUM / calculated NAV series"}
    # ISIN outranks the shapes below: an ISIN-bearing export is the bridge view.
    if has("isin") and (has("qty") or has("holding")):
        return {"type": "maia_full", "detail": "position export carrying ISIN"}
    if has("qty") and (has("exposure ($ usd)") or has("exp (fund ccy)")) \
            and (_norm("Accrued Total (Local)") in hdr or _norm("Accrued Total") in hdr):
        return {"type": "maia_priced", "detail": "priced view with accrued"}
    if has("grouping", "holding") and has("exp (fund ccy)"):
        return {"type": "maia_allocation", "detail": "allocation view (no accrued/price)"}
    if has("ticker", "qty", "strategy") and \
            (_norm("Exp Fund CCY") in hdr or _norm("Exp Inst CCY") in hdr):
        return {"type": "maia_positions",
                "detail": "positions/exposure shape (Instrument ID era) — no "
                          "Grouping, no accrued; not yet usable by the AUM recon"}
    return {"type": "unknown", "detail": f"unrecognised headers: {sorted(hdr)[:12]}"}
