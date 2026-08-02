"""Synthetic Maia-export fixtures — NO client data is ever committed.

Builders write openpyxl workbooks in the three real shapes (13-col
allocation, 34-col priced, 629-col-style full) with a tiny invented book:
two bonds (one EUR to exercise FX), one cash row per currency, one forward.
Numbers are chosen so totals are hand-checkable in the tests.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

BOOK = {
    # isin, ticker, desc, ccy, qty, clean_px, accrued_local, fx
    "bonds": [
        ("US0000000001", "T 4 2035", "US TREASURY 4% 2035", "USD",
         1_000_000, 99.0, 5_000.0, 1.0),
        ("DE0000000002", "DBR 2 2040", "BUND 2% 2040", "EUR",
         500_000, 90.0, 2_000.0, 1.10),
    ],
    "cash": [("USD", 100_000, 1.0), ("EUR", 50_000, 1.10)],
    "forwards": [("EURUSD Fwd 09/26", -3_000.0, "15/09/2026")],
}


def bond_exposure(qty, px, accrued_local, fx):
    """Dirty exposure in base ccy, as Maia reports it."""
    return (qty * px / 100 + accrued_local) * fx


def write_priced_view(path, valuation="31/07/2026 14:00:00"):
    """34-col-style PRICED shape: Grouping, Qty, Exposure, Last Px, Dirty
    price, Accrued Total (Local), NAV Contri — no ISIN."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    ws.append(["Grouping", "Ticker", "Description", "Currency", "Qty",
               "Exposure ($ USD)", "Last Px", "Dirty price",
               "Accrued Total (Local)", "NAV Contri Fund CCY (Alt 1)",
               "Settlement Date", "Last Valuation Date"])
    for isin, tkr, desc, ccy, qty, px, acc, fx in BOOK["bonds"]:
        exp = bond_exposure(qty, px, acc, fx)
        dirty_px = (exp / fx) / qty * 100
        ws.append(["Fixed Income", tkr, desc, ccy, qty, f"{exp:,.2f}",
                   px, dirty_px, f"{acc:,.2f}", "", "", valuation])
    for ccy, qty, fx in BOOK["cash"]:
        ws.append(["Currency", f"{ccy} Cash", f"{ccy} BALANCE", ccy,
                   qty, f"{qty * fx:,.2f}", "", "", "", "", "", valuation])
    for tkr, pnl, settle in BOOK["forwards"]:
        ws.append(["Currency", tkr, "FX FORWARD", "USD",
                   0, 0, "", "", "", f"{pnl:,.2f}", settle, valuation])
    wb.save(path)
    return path


def write_reference_view(path, valuation="31/07/2026 14:00:00",
                         drop_bonds=0):
    """Full-export-style shape carrying ISIN + Ticker (the bridge)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    ws.append(["Asset Class", "Ticker", "Description", "ISIN", "Currency",
               "Qty", "Exposure ($ USD)", "Last Px",
               "Last Valuation Date"])
    bonds = BOOK["bonds"][:len(BOOK["bonds"]) - drop_bonds]
    for isin, tkr, desc, ccy, qty, px, acc, fx in bonds:
        ws.append(["Government Bond", tkr, desc, isin, ccy, qty,
                   f"{bond_exposure(qty, px, acc, fx):,.2f}", px, valuation])
    for ccy, qty, fx in BOOK["cash"]:
        ws.append(["Cash", f"{ccy} Cash", f"{ccy} BALANCE", "", ccy,
                   qty, f"{qty * fx:,.2f}", "", valuation])
    wb.save(path)
    return path


def write_allocation_view(path, valuation="31/07/2026 14:00:00"):
    """13-col-style ALLOCATION shape: Holding + Exp (Fund CCY), no accrued,
    no price, no ISIN."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    ws.append(["Grouping", "Ticker", "Description", "Currency", "Holding",
               "Exp (Fund CCY)", "Last Valuation Date"])
    for isin, tkr, desc, ccy, qty, px, acc, fx in BOOK["bonds"]:
        ws.append(["Fixed Income", tkr, desc, ccy, qty,
                   f"{bond_exposure(qty, px, acc, fx):,.2f}", valuation])
    for ccy, qty, fx in BOOK["cash"]:
        ws.append(["Currency", f"{ccy} Cash", f"{ccy} BALANCE", ccy, qty,
                   f"{qty * fx:,.2f}", valuation])
    wb.save(path)
    return path


def admin_parsed(valuation_date="2026-07-31"):
    """A parsed administrator payload (nav_parser.parse_nav_report shape),
    built as a plain dict — the recon modules take the dict, not the file."""
    holdings = []
    for isin, tkr, desc, ccy, qty, px, acc, fx in BOOK["bonds"]:
        holdings.append({
            "isin": isin, "ticker": tkr, "description": desc,
            "currency": ccy, "par_amount": qty, "price": px,
            "market_value": qty * px / 100 * fx,       # clean, base ccy
            "accrued_income": acc * fx,                # base ccy
        })
    clean = sum(h["market_value"] for h in holdings)
    accrued = sum(h["accrued_income"] for h in holdings)
    cash = sum(q * fx for _, q, fx in BOOK["cash"])
    fwd = sum(p for _, p, _ in BOOK["forwards"])
    total = clean + accrued + cash + fwd
    return {
        "valuation_date": valuation_date,
        "valuation_time": "12:00",
        "base_currency": "USD",
        "portfolio_id": "gdbf",
        "fund_name": "Test Dynamic Bond Fund",
        "holdings": holdings,
        # Parser shape: {"rates": {ccy: UNITS PER BASE}} — attribution inverts.
        "fx_rates": {"rates": {"USD": 1.0, "EUR": 1.0 / 1.10}},
        "fx_forwards": [
            {"owner_type": "fund", "pnl": fwd,
             "legs": [{"settlement_date": "2026-09-15"}]},
        ],
        "summary": {
            "total_nav": total,
            "aum_breakdown": {
                "clean_bond_mv": clean, "accrued_income": accrued,
                "cash": cash, "fx_forward_pnl": fwd,
                "fx_forward_pnl_fund": fwd, "fx_forward_pnl_share_class": 0.0,
                "other_net": 0.0, "total_nav": total, "cash_residual": 0.0,
            },
        },
    }


@pytest.fixture
def priced_view(tmp_path):
    return str(write_priced_view(tmp_path / "priced.xlsx"))


@pytest.fixture
def reference_view(tmp_path):
    return str(write_reference_view(tmp_path / "ref.xlsx"))


@pytest.fixture
def allocation_view(tmp_path):
    return str(write_allocation_view(tmp_path / "alloc.xlsx"))
